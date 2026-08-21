from __future__ import annotations

from io import BytesIO
from types import SimpleNamespace
from zipfile import ZIP_DEFLATED, ZipFile

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from apps.api.admin.event_import import create_template
from apps.api.admin.middleware import AdminTraceMiddleware
from apps.api.admin.routes import public_router, router
from apps.api.admin.store import AdminStore
from apps.api.routes import qa as qa_routes


PNG_1X1 = b"\x89PNG\r\n\x1a\n"


def _client(tmp_path) -> TestClient:
    settings = SimpleNamespace(
        admin_sqlite_path=str(tmp_path / "admin.sqlite3"),
        admin_initialize_defaults=True,
        admin_jwt_secret="test-secret-that-is-long-enough-for-hs256",
        admin_access_token_minutes=30,
        admin_refresh_token_days=7,
        scene_assets_dir=str(tmp_path / "assets"),
        admin_event_import_root=str(tmp_path / "imports"),
    )
    app = FastAPI()
    app.state.settings = settings
    store = AdminStore(settings.admin_sqlite_path, True)
    store.save_record("exhibitions", {"id": "expo-test", "name": "测试展会", "status": "operating"}, "expo-test")
    app.state.admin_store = store
    app.add_middleware(AdminTraceMiddleware)
    app.include_router(router)
    app.include_router(public_router)
    app.include_router(qa_routes.router)
    return TestClient(app)


def _headers(client: TestClient) -> dict[str, str]:
    response = client.post("/api/v1/auth/login", json={"username": "admin", "password": "Admin@123456"})
    return {"Authorization": f"Bearer {response.json()['token']}"}


def _workbook() -> object:
    return load_workbook(BytesIO(create_template()))


def _append(workbook, sheet: str, values: dict[str, object]) -> None:
    worksheet = workbook[sheet]
    keys = [str(cell.value).split("（", 1)[0].lower() for cell in worksheet[1]]
    worksheet.append([values.get(key, "") for key in keys])


def test_template_marks_required_and_optional_columns() -> None:
    workbook = _workbook()
    headers = [str(cell.value) for cell in workbook["展商"][1]]
    assert "id（系统自动生成；更新时可填写）【系统生成】" in headers
    assert "category（行业分类）【选填】" in headers
    required = workbook["展商"][1][5]
    optional = workbook["展商"][1][7]
    assert required.comment is not None and "必填" in required.comment.text
    assert optional.comment is not None and "选填" in optional.comment.text


def test_new_records_receive_system_ids_and_relations_accept_human_keys(tmp_path) -> None:
    with _client(tmp_path) as client:
        headers = _headers(client)
        workbook = _workbook()
        _append(workbook, "展商", {"name": "自动编号展商", "booth_code": "B1-01", "status": "active"})
        _append(workbook, "场馆", {"exhibition_id": "expo-test", "name": "自动编号场馆", "address": "测试地址"})
        _append(workbook, "点位", {"exhibition_id": "expo-test", "venue_id": "自动编号场馆", "code": "P-A", "name": "自动入口", "type": "entrance"})
        _append(workbook, "点位", {"exhibition_id": "expo-test", "venue_id": "自动编号场馆", "code": "P-B", "name": "自动展台", "type": "booth"})
        _append(workbook, "路线", {"exhibition_id": "expo-test", "name": "自动路线", "type": "tour"})
        _append(workbook, "路线点位", {"route_id": "自动路线", "point_id": "P-A", "sort_order": 1})
        _append(workbook, "路线点位", {"route_id": "自动路线", "point_id": "P-B", "sort_order": 2})
        output = BytesIO()
        workbook.save(output)
        preview = client.post("/api/v1/admin/event/exhibitions/expo-test/import/preview", headers=headers, files={"file": ("auto-id.xlsx", output.getvalue())})
        assert preview.json()["canCommit"] is True
        client.post("/api/v1/admin/event/imports/commit", headers=headers, json={"batchId": preview.json()["batchId"]})
        exhibitors = client.app.state.admin_store.list_records("exhibitors", exhibition_id="expo-test")
        routes = client.app.state.admin_store.list_records("routes", exhibition_id="expo-test")
        assert any(item["name"] == "自动编号展商" and item["id"] and item["exhibitionId"] == "expo-test" for item in exhibitors)
        route = next(item for item in routes if item["name"] == "自动路线")
        assert len(route["pointIds"]) == 2


def test_template_preview_commit_and_idempotency(tmp_path) -> None:
    with _client(tmp_path) as client:
        headers = _headers(client)
        template = client.get("/api/v1/admin/event/exhibitions/expo-test/import-template", headers=headers)
        assert template.status_code == 200
        workbook = load_workbook(BytesIO(template.content))
        _append(workbook, "展商", {"id": "ex-1", "exhibition_id": "expo-test", "name": "展商一", "booth_code": "A1-01", "status": "active"})
        output = BytesIO()
        workbook.save(output)
        preview = client.post("/api/v1/admin/event/exhibitions/expo-test/import/preview", headers=headers, files={"file": ("data.xlsx", output.getvalue())})
        assert preview.status_code == 200
        payload = preview.json()
        assert payload["canCommit"] is True
        committed = client.post("/api/v1/admin/event/imports/commit", headers=headers, json={"batchId": payload["batchId"]})
        assert committed.status_code == 200
        repeated = client.post("/api/v1/admin/event/imports/commit", headers=headers, json={"batchId": payload["batchId"]})
        assert repeated.json()["idempotent"] is True
        assert client.app.state.admin_store.get_record("exhibitors", "ex-1")["name"] == "展商一"


def test_import_validates_relations_before_commit(tmp_path) -> None:
    with _client(tmp_path) as client:
        headers = _headers(client)
        workbook = _workbook()
        _append(workbook, "展品", {"id": "item-1", "exhibition_id": "expo-test", "exhibitor_id": "missing", "name": "无效展品"})
        output = BytesIO()
        workbook.save(output)
        response = client.post("/api/v1/admin/event/exhibitions/expo-test/import/preview", headers=headers, files={"file": ("data.xlsx", output.getvalue())})
        assert response.status_code == 200
        assert response.json()["canCommit"] is False
        assert any("展商" in item["message"] for item in response.json()["errors"])


def test_imports_knowledge_materials_and_public_qa_uses_documents(tmp_path) -> None:
    with _client(tmp_path) as client:
        headers = _headers(client)
        workbook = _workbook()
        _append(workbook, "知识库", {"id": "kb-expo", "exhibition_id": "expo-test", "name": "测试展会知识库"})
        _append(workbook, "知识文档", {"id": "doc-guide", "exhibition_id": "expo-test", "knowledge_base_id": "kb-expo", "title": "服务指南", "content": "服务中心位于一号馆入口西侧，提供咨询和失物招领。", "keywords": "服务中心,失物招领", "status": "published"})
        _append(workbook, "问答知识", {"id": "qa-guide", "exhibition_id": "expo-test", "question": "服务中心在哪里？", "keywords": "咨询台,失物招领", "answer": "服务中心位于一号馆入口西侧。", "status": "published"})
        output = BytesIO()
        workbook.save(output)
        preview = client.post("/api/v1/admin/event/exhibitions/expo-test/import/preview", headers=headers, files={"file": ("knowledge.xlsx", output.getvalue())})
        assert preview.status_code == 200
        assert preview.json()["canCommit"] is True
        committed = client.post("/api/v1/admin/event/imports/commit", headers=headers, json={"batchId": preview.json()["batchId"]})
        assert committed.status_code == 200
        assert client.app.state.admin_store.get_record("documents", "doc-guide")["content"].startswith("服务中心")
        assert client.app.state.admin_store.get_record("qa", "qa-guide")["status"] == "published"
        rag = client.post(
            "/exhibitions/expo-test/qa/query",
            json={"session_id": "demo", "turn_id": "turn-doc", "question": "一号馆入口西侧有什么服务？"},
        )
        assert rag.status_code == 200
        assert rag.json()["match_type"] == "rag"
        assert rag.json()["sources"][0]["title"] == "服务指南"


def test_zip_images_are_stored_and_route_points_are_ordered(tmp_path) -> None:
    with _client(tmp_path) as client:
        headers = _headers(client)
        workbook = _workbook()
        _append(workbook, "展商", {"id": "ex-1", "exhibition_id": "expo-test", "name": "展商一", "booth_code": "A1-01", "image_mode": "replace", "image_refs": "logo.png"})
        _append(workbook, "场馆", {"id": "venue-1", "exhibition_id": "expo-test", "name": "一号馆", "address": "地址"})
        _append(workbook, "点位", {"id": "point-1", "exhibition_id": "expo-test", "venue_id": "venue-1", "code": "P1", "name": "入口", "type": "entrance"})
        _append(workbook, "点位", {"id": "point-2", "exhibition_id": "expo-test", "venue_id": "venue-1", "code": "P2", "name": "展台", "type": "booth"})
        _append(workbook, "路线", {"id": "route-1", "exhibition_id": "expo-test", "name": "入口路线", "type": "tour"})
        _append(workbook, "路线点位", {"route_id": "route-1", "point_id": "point-2", "sort_order": 2})
        _append(workbook, "路线点位", {"route_id": "route-1", "point_id": "point-1", "sort_order": 1})
        output = BytesIO()
        workbook.save(output)
        package = BytesIO()
        with ZipFile(package, "w", ZIP_DEFLATED) as archive:
            archive.writestr("data.xlsx", output.getvalue())
            archive.writestr("images/logo.png", PNG_1X1)
        preview = client.post("/api/v1/admin/event/exhibitions/expo-test/import/preview", headers=headers, files={"file": ("bundle.zip", package.getvalue(), "application/zip")})
        assert preview.json()["canCommit"] is True
        assert client.post("/api/v1/admin/event/imports/commit", headers=headers, json={"batchId": preview.json()["batchId"]}).status_code == 200
        assert client.app.state.admin_store.get_record("routes", "route-1")["pointIds"] == ["point-1", "point-2"]
        assert client.app.state.admin_store.get_record("exhibitors", "ex-1")["imageUrls"]
        config = client.get("/exhibitions/expo-test/digital-human-config")
        assert config.status_code == 200
        assert "入口路线" in config.json()["keywords"]["navigation"]
        navigation = client.post("/exhibitions/expo-test/navigation/query", json={"text": "入口路线怎么走", "session_id": "import-nav"})
        assert navigation.status_code == 200
        assert navigation.json()["matched"] is True
        assert navigation.json()["route_id"] == "route-1"
