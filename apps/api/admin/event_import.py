from __future__ import annotations

"""Excel import/export helpers for exhibition operational data.

The importer intentionally keeps the workbook human-editable: one workbook has
one sheet per resource, IDs are stable, and relationship rows are explicit in
the ``路线点位`` sheet.  Images are kept out of the workbook and are supplied
as files in a sibling ``images/`` directory inside the uploaded ZIP package.
"""

import io
import re
import uuid
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path, PurePosixPath
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


TEMPLATE_VERSION = "1.1"
MAX_PACKAGE_BYTES = 100 * 1024 * 1024
MAX_WORKBOOK_ROWS = 20_000
MAX_IMAGE_BYTES = 10 * 1024 * 1024
ALLOWED_IMAGE_TYPES = {
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".webp": "image/webp",
    ".gif": "image/gif",
    ".svg": "image/svg+xml",
}


@dataclass(frozen=True)
class SheetSpec:
    title: str
    kind: str | None
    fields: tuple[tuple[str, str, bool, str], ...]


COMMON_FIELDS: tuple[tuple[str, str, bool, str], ...] = (
    ("id（系统自动生成；更新时可填写）", "id", False, "新增记录留空，系统自动生成；更新已有记录时可填写原 ID"),
    ("exhibition_id（当前展会自动填充）", "exhibitionId", False, "系统根据当前展会详情自动填充，无需手工填写"),
    ("image_mode（图片策略）", "imageMode", False, "keep 保留；replace 替换；clear 清空"),
    ("image_refs（附件文件名）", "imageRefs", False, "ZIP 内 images/ 下文件名，每行一个"),
    ("image_urls（已有URL）", "imageUrls", False, "已有图片 URL，每行一个"),
)


SHEETS: tuple[SheetSpec, ...] = (
    SheetSpec("展商", "exhibitors", COMMON_FIELDS + (
        ("name（展商名称）", "name", True, "展商全称"),
        ("booth_code（展位号）", "boothCode", True, "如 A1-08"),
        ("category（行业分类）", "category", False, "行业或主题分类"),
        ("contact（联系人）", "contact", False, "现场联系人"),
        ("phone（联系电话）", "phone", False, "手机号或座机"),
        ("status（状态）", "status", False, "pending / active / inactive"),
        ("description（简介）", "description", False, "支持 Markdown"),
        ("introduction_keywords（介绍关键词）", "introductionKeywords", False, "每行一个"),
        ("aliases（别名）", "aliases", False, "每行一个"),
        ("fuzzy_match（模糊匹配）", "fuzzyMatch", False, "TRUE / FALSE"),
        ("spoken_text（播报文本）", "spokenText", False, "数字人播报文本"),
    )),
    SheetSpec("展品", "exhibits", COMMON_FIELDS + (
        ("exhibitor_id（展商ID/名称/展位号）", "exhibitorId", True, "可填写展商系统 ID、展商名称或展位号；系统会自动解析"),
        ("name（展品名称）", "name", True, "展品名称"),
        ("category（类别）", "category", False, "展品类别"),
        ("model_no（型号）", "modelNo", False, "产品型号"),
        ("description（说明）", "description", False, "支持 Markdown"),
        ("status（状态）", "status", False, "draft / published"),
        ("introduction_keywords（介绍关键词）", "introductionKeywords", False, "每行一个"),
        ("aliases（别名）", "aliases", False, "每行一个"),
        ("fuzzy_match（模糊匹配）", "fuzzyMatch", False, "TRUE / FALSE"),
        ("spoken_text（播报文本）", "spokenText", False, "数字人播报文本"),
    )),
    SheetSpec("场馆", "venues", COMMON_FIELDS + (
        ("name（场馆名称）", "name", True, "场馆名称"),
        ("address（地址）", "address", True, "场馆地址"),
        ("description（说明）", "description", False, "场馆说明"),
        ("status（状态）", "status", False, "draft / active / inactive"),
        ("introduction_keywords（介绍关键词）", "introductionKeywords", False, "每行一个"),
        ("aliases（别名）", "aliases", False, "每行一个"),
        ("fuzzy_match（模糊匹配）", "fuzzyMatch", False, "TRUE / FALSE"),
        ("spoken_text（播报文本）", "spokenText", False, "数字人播报文本"),
    )),
    SheetSpec("点位", "points", COMMON_FIELDS + (
        ("venue_id（场馆ID/名称）", "venueId", True, "可填写场馆系统 ID 或场馆名称；系统会自动解析"),
        ("code（点位编码）", "code", True, "现场点位唯一编码"),
        ("name（点位名称）", "name", True, "点位名称"),
        ("type（点位类型）", "type", True, "entrance / booth / forum / facility / service / other"),
        ("floor（楼层）", "floor", False, "如 1F"),
        ("x（地图X）", "x", False, "地图坐标"),
        ("y（地图Y）", "y", False, "地图坐标"),
        ("exhibitor_id（展商ID/名称/展位号）", "exhibitorId", False, "可填写展商系统 ID、展商名称或展位号"),
        ("exhibit_id（展品ID/名称）", "exhibitId", False, "可填写展品系统 ID 或展品名称"),
        ("description（说明）", "description", False, "点位说明"),
        ("status（状态）", "status", False, "active / inactive"),
        ("introduction_keywords（介绍关键词）", "introductionKeywords", False, "每行一个"),
        ("aliases（别名）", "aliases", False, "每行一个"),
        ("fuzzy_match（模糊匹配）", "fuzzyMatch", False, "TRUE / FALSE"),
        ("spoken_text（播报文本）", "spokenText", False, "数字人播报文本"),
    )),
    SheetSpec("路线", "routes", COMMON_FIELDS + (
        ("name（路线名称）", "name", True, "路线名称"),
        ("type（路线类型）", "type", True, "navigation / tour / emergency"),
        ("keywords（寻路关键词）", "keywords", False, "每行一个"),
        ("aliases（目的地别名）", "aliases", False, "每行一个"),
        ("fuzzy_match（模糊匹配）", "fuzzyMatch", False, "TRUE / FALSE"),
        ("directions（路段指引）", "directions", False, "每行一段；为空时自动生成"),
        ("spoken_text（播报文本）", "spokenText", False, "数字人播报文本"),
        ("estimated_minutes（预计分钟）", "estimatedMinutes", False, "整数"),
        ("description（说明）", "description", False, "路线说明"),
        ("status（状态）", "status", False, "draft / published"),
    )),
    SheetSpec("路线点位", "route_points", (
        ("route_id（路线ID/路线名称）", "routeId", True, "可填写路线系统 ID 或路线名称；系统会自动解析"),
        ("point_id（点位ID/编码）", "pointId", True, "可填写点位系统 ID、点位编码或点位名称；系统会自动解析"),
        ("sort_order（顺序）", "sortOrder", True, "从 1 开始"),
    )),
    SheetSpec("活动排期", "schedules", COMMON_FIELDS + (
        ("venue_id（场馆ID/名称）", "venueId", False, "可填写场馆系统 ID 或场馆名称"),
        ("point_id（点位ID/编码）", "pointId", False, "可填写点位系统 ID、编码或名称"),
        ("title（活动标题）", "title", True, "活动标题"),
        ("type（活动类型）", "type", False, "论坛 / 演示 / 签到等"),
        ("start_at（开始时间）", "startAt", True, "YYYY-MM-DD HH:mm"),
        ("end_at（结束时间）", "endAt", True, "YYYY-MM-DD HH:mm"),
        ("location（活动地点）", "location", False, "展示用地点"),
        ("speaker（主讲人）", "speaker", False, "主讲人或单位"),
        ("description（说明）", "description", False, "活动说明"),
        ("status（状态）", "status", False, "draft / scheduled / finished / cancelled"),
    )),
    SheetSpec("应急播报", "broadcasts", COMMON_FIELDS + (
        ("title（标题）", "title", True, "播报标题"),
        ("content（内容）", "content", True, "播报正文"),
        ("priority（优先级）", "priority", False, "low / normal / high / urgent"),
        ("target_terminals（目标终端）", "targetTerminals", False, "全部终端或终端 ID"),
        ("effective_at（生效时间）", "effectiveAt", False, "YYYY-MM-DD HH:mm"),
        ("status（状态）", "status", False, "draft / active / ended"),
    )),
    SheetSpec("知识库", "knowledge_bases", (
        ("id（系统自动生成；更新时可填写）", "id", False, "新增记录留空，系统自动生成；更新已有记录时可填写原 ID"),
        ("exhibition_id（当前展会自动填充）", "exhibitionId", False, "系统根据当前展会详情自动填充，无需手工填写"),
        ("name（知识库名称）", "name", True, "知识库展示名称"),
        ("local_knowledge_base_id（本地知识库ID）", "localKnowledgeBaseId", False, "已建立本地向量库时填写；为空时使用记录 ID"),
        ("description（说明）", "description", False, "知识库用途说明"),
        ("status（状态）", "status", False, "active / inactive"),
    )),
    SheetSpec("知识文档", "documents", (
        ("id（系统自动生成；更新时可填写）", "id", False, "新增记录留空，系统自动生成；更新已有记录时可填写原 ID"),
        ("exhibition_id（当前展会自动填充）", "exhibitionId", False, "系统根据当前展会详情自动填充，无需手工填写"),
        ("knowledge_base_id（知识库ID）", "knowledgeBaseId", True, "关联知识库表的 ID"),
        ("title（标题）", "title", True, "文档标题"),
        ("content（正文）", "content", True, "可直接检索的正文；支持 Markdown"),
        ("keywords（关键词）", "keywords", False, "每行一个"),
        ("category（分类）", "category", False, "如展会概览、服务指南、展商资料"),
        ("source_url（来源URL）", "sourceUrl", False, "可选来源地址"),
        ("status（状态）", "status", False, "published / draft / archived"),
    )),
    SheetSpec("问答知识", "qa", (
        ("id（系统自动生成；更新时可填写）", "id", False, "新增记录留空，系统自动生成；更新已有记录时可填写原 ID"),
        ("exhibition_id（当前展会自动填充）", "exhibitionId", False, "系统根据当前展会详情自动填充，无需手工填写"),
        ("question（标准问题）", "question", True, "用户可能提出的问题"),
        ("keywords（匹配关键词）", "keywords", False, "每行一个"),
        ("answer（官方答案）", "answer", True, "审核后直接播报的答案"),
        ("category（分类）", "category", False, "如展会、交通、服务、展商"),
        ("status（状态）", "status", False, "published / draft / pending_review / archived"),
        ("version（版本）", "version", False, "整数版本号"),
        ("creator（创建人）", "creator", False, "创建人"),
    )),
)


FIELD_ALIASES = {
    "id": "id", "记录id": "id", "记录ID": "id",
    "exhibitionid": "exhibitionId", "展会id": "exhibitionId", "展会ID": "exhibitionId",
    "image_mode": "imageMode", "图片策略": "imageMode", "imagemode": "imageMode",
    "image_refs": "imageRefs", "附件文件名": "imageRefs", "imagerefs": "imageRefs",
    "image_urls": "imageUrls", "已有url": "imageUrls", "imageurls": "imageUrls",
}

for _spec in SHEETS:
    for _label, _key, _required, _description in _spec.fields:
        _plain_label = re.sub(r"（.*?）|\(.*?\)", "", _label).strip().lower()
        FIELD_ALIASES.setdefault(_plain_label, _key)
        FIELD_ALIASES.setdefault(_key.lower(), _key)
        FIELD_ALIASES.setdefault(re.sub(r"(?<!^)(?=[A-Z])", "_", _key).lower(), _key)


def _header_key(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"(?:【|\[)(?:必填|选填|系统生成|系统填充)(?:】|\])", "", text)
    text = re.sub(r"（.*?）|\(.*?\)", "", text).strip().lower()
    text = re.sub(r"[\s\-]+", "_", text)
    return FIELD_ALIASES.get(text, text)


def _split_values(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, (list, tuple)):
        values = value
    else:
        values = re.split(r"[\n,，、;；]+", str(value))
    return list(dict.fromkeys(str(item).strip() for item in values if str(item).strip()))


def _scalar(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ") if isinstance(value, datetime) else value.isoformat()
    return value


def _bool(value: Any, default: bool = True) -> bool:
    if value is None or str(value).strip() == "":
        return default
    return str(value).strip().lower() in {"1", "true", "yes", "y", "是", "启用"}


def _safe_member_name(name: str) -> str:
    path = PurePosixPath(name)
    if path.is_absolute() or ".." in path.parts:
        raise ValueError("ZIP 包含非法路径")
    return str(path)


def _image_mime(name: str, content: bytes) -> str | None:
    extension = Path(name).suffix.lower()
    mime = ALLOWED_IMAGE_TYPES.get(extension)
    if not mime:
        return None
    if mime == "image/png" and not content.startswith(b"\x89PNG\r\n\x1a\n"):
        return None
    if mime == "image/jpeg" and not content.startswith(b"\xff\xd8\xff"):
        return None
    if mime == "image/gif" and not content.startswith((b"GIF87a", b"GIF89a")):
        return None
    if mime == "image/webp" and not (len(content) >= 12 and content[:4] == b"RIFF" and content[8:12] == b"WEBP"):
        return None
    if mime == "image/svg+xml" and b"<svg" not in content[:4096].lower():
        return None
    return mime


def extract_package(payload: bytes, filename: str) -> tuple[bytes, dict[str, tuple[bytes, str]]]:
    if len(payload) > MAX_PACKAGE_BYTES:
        raise ValueError("导入文件不能超过 100MB")
    if Path(filename).suffix.lower() != ".zip":
        return payload, {}
    try:
        archive = zipfile.ZipFile(io.BytesIO(payload))
    except zipfile.BadZipFile as exc:
        raise ValueError("图片附件包不是有效 ZIP 文件") from exc
    workbook: bytes | None = None
    images: dict[str, tuple[bytes, str]] = {}
    expanded_bytes = 0
    for info in archive.infolist():
        member = _safe_member_name(info.filename)
        if info.is_dir():
            continue
        expanded_bytes += int(info.file_size or 0)
        if expanded_bytes > MAX_PACKAGE_BYTES:
            raise ValueError("ZIP 解压后的总大小不能超过 100MB")
        if int(info.file_size or 0) > MAX_IMAGE_BYTES and member.lower().startswith("images/"):
            raise ValueError(f"图片 {PurePosixPath(member).name} 超过 10MB")
        content = archive.read(info)
        if member.lower() in {"data.xlsx", "workbook.xlsx", "导入数据.xlsx"}:
            workbook = content
            continue
        if member.lower().startswith("images/"):
            filename_only = PurePosixPath(member).name
            mime = _image_mime(filename_only, content)
            if mime:
                if len(content) > MAX_IMAGE_BYTES:
                    raise ValueError(f"图片 {filename_only} 超过 10MB")
                images[filename_only] = (content, mime)
    if workbook is None:
        raise ValueError("ZIP 包内必须包含 data.xlsx")
    return workbook, images


def create_template() -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="0F766E")
    required_header_fill = PatternFill("solid", fgColor="9F1239")
    generated_header_fill = PatternFill("solid", fgColor="475569")
    context_header_fill = PatternFill("solid", fgColor="B45309")
    header_font = Font(color="FFFFFF", bold=True)
    for spec in SHEETS:
        sheet = workbook.create_sheet(spec.title)
        if spec.kind is None:
            continue
        def marker(field: tuple[str, str, bool, str]) -> str:
            if field[1] == "id":
                return "系统生成"
            if field[1] == "exhibitionId":
                return "系统填充"
            return "必填" if field[2] else "选填"

        sheet.append([f"{field[0]}【{marker(field)}】" for field in spec.fields])
        for index, cell in enumerate(sheet[1]):
            field = spec.fields[index]
            cell.fill = context_header_fill if field[1] == "exhibitionId" else (generated_header_fill if field[1] == "id" else (required_header_fill if field[2] else header_fill))
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.comment = Comment(
                f"字段属性：{marker(field)}。{field[3]}",
                "OpenTalking",
            )
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{chr(64 + min(len(spec.fields), 26))}2"
        for index, field in enumerate(spec.fields, 1):
            sheet.column_dimensions[chr(64 + index) if index <= 26 else "A"].width = min(max(len(field[0]) + 4, 14), 30)
            sheet.cell(2, index).comment = None
        if spec.title in {"展商", "展品", "场馆", "点位", "路线", "活动排期", "应急播报"}:
            dv = DataValidation(type="list", formula1='"keep,replace,clear"', allow_blank=True)
            sheet.add_data_validation(dv)
            image_mode_column = next(index for index, field in enumerate(spec.fields, 1) if field[1] == "imageMode")
            dv.add(f"{chr(64 + image_mode_column)}2:{chr(64 + image_mode_column)}{MAX_WORKBOOK_ROWS}")

    instructions = workbook.create_sheet("说明", 0)
    instructions.append(["展会运营数据导入模板", f"模板版本 {TEMPLATE_VERSION}"])
    instructions.append(["使用方式", "1. 新增记录时 ID 留空，由系统自动生成；更新已有记录时才填写原 ID。2. 路线点位在‘路线点位’表维护顺序；3. 图片请与 data.xlsx 一起放进 ZIP 的 images/ 目录；4. 知识库、知识文档、问答知识会随展会一起导入；5. 上传后先预览校验，再确认提交。"])
    instructions.append(["字段标识", "表头末尾的【必填】必须填写；【选填】可留空；【系统生成】由系统自动生成，更新时才填写；【系统填充】由当前展会详情自动填充。玫红色表头为必填字段，青绿色表头为选填字段，灰色表头为系统生成字段，棕色表头为系统填充字段；本页下方附有字段说明和数据字典。"])
    instructions.append(["图片策略", "keep=保留原图片；replace=使用 image_refs/image_urls 替换；clear=清空图片。image_refs 每行一个附件文件名。"])
    instructions.append(["更新规则", "同一 kind + id 会更新；模板中未出现的记录不会删除；任何错误都会阻止提交。"])
    instructions.column_dimensions["A"].width = 18
    instructions.column_dimensions["B"].width = 120
    instructions["A1"].font = Font(size=16, bold=True, color="0F766E")
    instructions["B1"].font = Font(bold=True)
    for row in instructions.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    instructions.append([])
    instructions.append(["字段说明", "工作表", "字段", "内部字段", "属性", "说明"])
    field_header_row = instructions.max_row
    for cell in instructions[field_header_row]:
        cell.fill = header_fill
        cell.font = header_font
    for spec in SHEETS:
        for label, key, required, description in spec.fields:
            attribute = "系统生成" if key == "id" else ("系统填充" if key == "exhibitionId" else ("必填" if required else "选填"))
            instructions.append(["", spec.title, label, key, attribute, description])

    instructions.append([])
    instructions.append(["数据字典", "字典", "可选值", "说明"])
    dictionary_header_row = instructions.max_row
    for cell in instructions[dictionary_header_row]:
        cell.fill = header_fill
        cell.font = header_font
    dictionaries = [
        ("图片策略", "keep", "保留原图片"), ("图片策略", "replace", "替换为导入图片"), ("图片策略", "clear", "清空图片"),
        ("点位类型", "entrance / booth / forum / facility / service / other", "点位类型"),
        ("路线类型", "navigation / tour / emergency", "路线类型"),
        ("展商状态", "pending / active / inactive", "展商状态"),
        ("展品状态", "draft / published", "展品状态"),
        ("排期状态", "draft / scheduled / finished / cancelled", "活动排期状态"),
        ("播报优先级", "low / normal / high / urgent", "应急播报优先级"),
        ("播报状态", "draft / active / ended", "应急播报状态"),
        ("知识库状态", "active / inactive", "知识库状态"),
        ("知识文档状态", "published / draft / archived", "知识文档状态"),
        ("问答状态", "published / draft / pending_review / archived", "官方问答审核状态"),
    ]
    for row in dictionaries:
        instructions.append(["", *row])
    for col, width in {"A": 18, "B": 18, "C": 34, "D": 24, "E": 10, "F": 80}.items():
        instructions.column_dimensions[col].width = width
    instructions.freeze_panes = "A2"
    for row in instructions.iter_rows(min_row=field_header_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _spec_by_title(title: str) -> SheetSpec | None:
    return next((spec for spec in SHEETS if spec.title == title), None)


def parse_workbook(workbook_bytes: bytes, images: dict[str, tuple[bytes, str]]) -> tuple[dict[str, list[dict[str, Any]]], list[dict[str, Any]]]:
    try:
        workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=True, read_only=True)
    except Exception as exc:  # openpyxl raises several file-format-specific exceptions
        raise ValueError("无法读取 Excel 文件，请使用模板生成的 .xlsx 文件") from exc
    records: dict[str, list[dict[str, Any]]] = {}
    errors: list[dict[str, Any]] = []
    for title in workbook.sheetnames:
        spec = _spec_by_title(title)
        if not spec or not spec.kind:
            continue
        sheet = workbook[title]
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration:
            continue
        key_by_column = [_header_key(value) for value in headers]
        expected = {key for _, key, _, _ in spec.fields}
        parsed: list[dict[str, Any]] = []
        for row_number, values in enumerate(rows, 2):
            if not any(value not in (None, "") for value in values):
                continue
            if row_number > MAX_WORKBOOK_ROWS:
                errors.append({"sheet": title, "row": row_number, "message": f"单个工作表不能超过 {MAX_WORKBOOK_ROWS} 行"})
                break
            raw = {key: _scalar(values[index]) for index, key in enumerate(key_by_column) if key and index < len(values) and values[index] not in (None, "")}
            unknown = set(raw) - expected
            if unknown:
                errors.append({"sheet": title, "row": row_number, "message": f"未知字段：{', '.join(sorted(unknown))}"})
            if spec.kind == "route_points":
                item = {"routeId": str(raw.get("routeId") or "").strip(), "pointId": str(raw.get("pointId") or "").strip(), "sortOrder": raw.get("sortOrder")}
            else:
                item = dict(raw)
                if not str(item.get("id") or "").strip():
                    item["id"] = f"{spec.kind.rstrip('_s')}-{uuid.uuid4().hex[:12]}"
                for field in ("keywords", "aliases", "directions", "imageRefs", "imageUrls", "introductionKeywords"):
                    if field in item:
                        item[field] = _split_values(item[field])
                for field in ("fuzzyMatch",):
                    if field in item:
                        item[field] = _bool(item[field])
                for field in ("x", "y", "estimatedMinutes"):
                    if field in item and item[field] not in (None, ""):
                        try:
                            item[field] = float(item[field]) if field in {"x", "y"} else int(float(item[field]))
                        except (TypeError, ValueError):
                            errors.append({"sheet": title, "row": row_number, "field": field, "message": "必须是数字"})
            item["_sheet"] = title
            item["_row"] = row_number
            for raw_filename in item.get("imageRefs", []):
                filename = str(raw_filename).strip()
                if Path(filename).name != filename:
                    errors.append({"sheet": title, "row": row_number, "field": "image_refs", "message": "附件文件名不能包含目录路径"})
                    continue
                if filename not in images:
                    errors.append({"sheet": title, "row": row_number, "field": "image_refs", "message": f"附件不存在：{filename}"})
            records.setdefault(spec.kind, []).append(item)
    return records, errors


def normalized_image_urls(item: dict[str, Any]) -> list[str]:
    urls = [str(value).strip() for value in item.get("imageUrls", []) if str(value).strip()]
    invalid = [url for url in urls if not (url.startswith("/") or url.startswith("http://") or url.startswith("https://"))]
    if invalid:
        raise ValueError(f"图片 URL 格式不正确：{invalid[0]}")
    return list(dict.fromkeys(urls))


def public_preview(batch_id: str, exhibition_id: str, filename: str, records: dict[str, list[dict[str, Any]]], errors: list[dict[str, Any]], warnings: list[dict[str, Any]]) -> dict[str, Any]:
    summary: dict[str, dict[str, int]] = {}
    conflicts: list[dict[str, Any]] = []
    for kind, items in records.items():
        summary[kind] = {"total": len(items), "errors": 0, "warnings": 0, "creates": 0, "updates": 0}
        for item in items:
            item_id = str(item.get("id") or item.get("routeId") or "")
            if item_id:
                conflicts.append({"kind": kind, "id": item_id, "action": "update"})
    for error in errors:
        kind = next((spec.kind for spec in SHEETS if spec.title == error.get("sheet")), None)
        if kind and kind in summary:
            summary[kind]["errors"] += 1
    return {"batchId": batch_id, "exhibitionId": exhibition_id, "filename": filename, "summary": summary, "conflicts": conflicts, "errors": errors, "warnings": warnings, "canCommit": not errors}
